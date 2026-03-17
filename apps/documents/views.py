from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from .models import ConfigurationDocument, CertificatScolarite
from apps.academic.models import AnneeScolaire, SalleClasse, Niveau
from apps.students.models import Eleve, Inscription
from apps.authentication.models import CustomUser


def role_requis(*roles):
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('login')
            if not (request.user.is_superuser or request.user.role in roles):
                messages.error(request, "Acces refuse.")
                return redirect('dashboard')
            return view_func(request, *args, **kwargs)
        wrapper.__name__ = view_func.__name__
        return wrapper
    return decorator


# ── CONFIGURATION DOCUMENTS ───────────────────────────────────────────────────

@login_required
@role_requis('DIRECTEUR', 'CENSEUR', 'SECRETAIRE')
def liste_documents(request):
    annee = AnneeScolaire.active()
    config = ConfigurationDocument.get()
    certificats = CertificatScolarite.objects.select_related(
        'eleve', 'annee', 'delivre_par'
    ).order_by('-date_delivrance')[:20]

    return render(request, 'documents/liste_documents.html', {
        'config': config,
        'certificats': certificats,
        'annee': annee,
    })


@login_required
@role_requis('DIRECTEUR')
def modifier_config_documents(request):
    config = ConfigurationDocument.get()
    if request.method == 'POST':
        config.nom_ecole = request.POST.get('nom_ecole', '').strip()
        config.slogan = request.POST.get('slogan', '').strip()
        config.adresse = request.POST.get('adresse', '').strip()
        config.telephone = request.POST.get('telephone', '').strip()
        config.email = request.POST.get('email', '').strip()
        config.site_web = request.POST.get('site_web', '').strip()
        config.pied_page = request.POST.get('pied_page', '').strip()
        config.ministre_tutelle = request.POST.get('ministre_tutelle', '').strip()
        config.devise_nationale = request.POST.get('devise_nationale', '').strip()
        if request.FILES.get('logo'):
            config.logo = request.FILES['logo']
        if request.FILES.get('signature_directeur'):
            config.signature_directeur = request.FILES['signature_directeur']
        if request.FILES.get('cachet'):
            config.cachet = request.FILES['cachet']
        config.save()
        messages.success(request, "Configuration documents mise a jour.")
        return redirect('liste_documents')
    return render(request, 'documents/modifier_config.html', {'config': config})


# ── CERTIFICATS DE SCOLARITE ──────────────────────────────────────────────────

@login_required
@role_requis('DIRECTEUR', 'CENSEUR', 'SECRETAIRE')
def generer_certificat(request, eleve_pk):
    eleve = get_object_or_404(Eleve, pk=eleve_pk)
    annee = AnneeScolaire.active()

    if not annee:
        messages.error(request, "Aucune annee active.")
        return redirect('liste_documents')

    motif = request.GET.get('motif', 'Toutes fins utiles')

    from django.utils import timezone
    numero = f"CERT-{timezone.now().strftime('%Y%m%d%H%M%S')}"

    certificat = CertificatScolarite.objects.create(
        eleve=eleve,
        annee=annee,
        numero=numero,
        delivre_par=request.user,
        motif=motif,
    )

    config = ConfigurationDocument.get()
    return _pdf_certificat(certificat, config)


@login_required
@role_requis('DIRECTEUR', 'CENSEUR', 'SECRETAIRE')
def reimprimer_certificat(request, pk):
    certificat = get_object_or_404(CertificatScolarite, pk=pk)
    config = ConfigurationDocument.get()
    return _pdf_certificat(certificat, config)


def _pdf_certificat(certificat, config):
    import io
    import qrcode
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer,
        Table, TableStyle, HRFlowable
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT
    from reportlab.pdfgen import canvas as pdfcanvas

    eleve = certificat.eleve
    annee = certificat.annee
    inscription = eleve.inscriptions.filter(
        annee=annee, statut='ACTIVE'
    ).select_related('salle__niveau').first()

    BLEU_FONCE = colors.HexColor('#1E3A8A')
    BLEU_CLAIR = colors.HexColor('#DBEAFE')
    BLEU_MED   = colors.HexColor('#3B82F6')
    OR         = colors.HexColor('#D97706')
    GRIS_CLAIR = colors.HexColor('#F8FAFC')
    BORDURE    = colors.HexColor('#E2E8F0')

    styles = getSampleStyleSheet()

    def s(name, **kw):
        return ParagraphStyle(name, parent=styles['Normal'], **kw)

    # ── QR CODE ───────────────────────────────────────────────────────────────
    qr_data = (
        f"GESTO|CERT|{certificat.numero}|"
        f"{eleve.matricule}|{annee.nom}"
    )
    qr = qrcode.QRCode(version=1, box_size=4, border=2)
    qr.add_data(qr_data)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="#1E3A8A", back_color="white")
    qr_buf = io.BytesIO()
    qr_img.save(qr_buf, format='PNG')
    qr_buf.seek(0)

    from reportlab.platypus import Image as RLImage
    qr_image = RLImage(qr_buf, width=2.2*cm, height=2.2*cm)

    # ── FILIGRANE ─────────────────────────────────────────────────────────────
    def filigrane(canvas_obj, doc):
        canvas_obj.saveState()
        canvas_obj.setFont('Helvetica-Bold', 55)
        canvas_obj.setFillColorRGB(0.93, 0.95, 0.98)
        canvas_obj.translate(A4[0]/2, A4[1]/2)
        canvas_obj.rotate(38)
        nom_court = config.nom[:18] if config.nom else 'ECOLE'
        canvas_obj.drawCentredString(0, 0, nom_court)
        canvas_obj.restoreState()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )

    elements = []

    # ── BANDE BLEUE EN-TETE ───────────────────────────────────────────────────
    entete = Table([[
        Paragraph(
            f"République Togolaise<br/>"
            f"<font size=8><b>Travail – Liberté – Patrie</b></font>",
            s('e1', fontSize=8, textColor=colors.white,
              alignment=TA_CENTER)
        ),
        Paragraph(
            f"<b><font size=13>{config.nom}</font></b><br/>"
            f"<font size=8>{config.adresse or ''}</font><br/>"
            f"<font size=8>Tél : {config.telephone or ''}</font>",
            s('e2', fontSize=9, textColor=colors.white,
              alignment=TA_CENTER)
        ),
        Paragraph(
            f"{config.ministre_tutelle or 'Ministère des Enseignements'}<br/>"
            f"<font size=8>Direction Régionale</font>",
            s('e3', fontSize=8, textColor=colors.white,
              alignment=TA_CENTER)
        ),
    ]], colWidths=[5.5*cm, 7*cm, 5.5*cm])
    entete.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), BLEU_FONCE),
        ('TOPPADDING', (0,0), (-1,-1), 10),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ROUNDEDCORNERS', [6]),
    ]))
    elements.append(entete)
    elements.append(Spacer(1, 0.5*cm))

    # ── TITRE ─────────────────────────────────────────────────────────────────
    titre_box = Table([[
        Paragraph(
            "CERTIFICAT DE SCOLARITÉ",
            s('titre', fontSize=16, fontName='Helvetica-Bold',
              textColor=colors.white, alignment=TA_CENTER)
        )
    ]], colWidths=[19*cm])
    titre_box.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), BLEU_MED),
        ('TOPPADDING', (0,0), (-1,-1), 10),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ('ROUNDEDCORNERS', [5]),
    ]))
    elements.append(titre_box)
    elements.append(Paragraph(
        f"Année scolaire {annee.nom}",
        s('as', fontSize=10, alignment=TA_CENTER,
          textColor=BLEU_FONCE, spaceBefore=4, spaceAfter=12)
    ))

    # ── CORPS ─────────────────────────────────────────────────────────────────
    salle_nom = inscription.salle.nom if inscription else 'Non inscrit'
    niveau_nom = inscription.salle.niveau.nom if inscription else ''
    sexe = 'la nommée' if eleve.sexe == 'F' else 'le nommé'
    article = 'Elle' if eleve.sexe == 'F' else 'Il'
    inscrit = 'inscrite' if eleve.sexe == 'F' else 'inscrit'

    elements.append(Paragraph(
        f"Je soussigné(e), Directeur(rice) de l'établissement "
        f"<b>{config.nom}</b>, certifie que {sexe} :",
        s('corps', fontSize=11, leading=18, alignment=TA_JUSTIFY,
          spaceAfter=10)
    ))

    # Fiche identité élève
    fiche = Table([
        [
            Paragraph('<b>Nom et Prénoms</b>',
                      s('fl', fontSize=10, textColor=BLEU_FONCE)),
            Paragraph(f'<b>{eleve.nom_complet}</b>',
                      s('fv', fontSize=11, fontName='Helvetica-Bold')),
        ],
        [
            Paragraph('<b>Matricule</b>',
                      s('fl', fontSize=10, textColor=BLEU_FONCE)),
            Paragraph(eleve.matricule, s('fv', fontSize=10)),
        ],
        [
            Paragraph('<b>Date de naissance</b>',
                      s('fl', fontSize=10, textColor=BLEU_FONCE)),
            Paragraph(
                eleve.date_naissance.strftime('%d/%m/%Y')
                if eleve.date_naissance else '—',
                s('fv', fontSize=10)
            ),
        ],
        [
            Paragraph('<b>Lieu de naissance</b>',
                      s('fl', fontSize=10, textColor=BLEU_FONCE)),
            Paragraph(
                eleve.lieu_naissance or '—',
                s('fv', fontSize=10)
            ),
        ],
        [
            Paragraph('<b>Classe</b>',
                      s('fl', fontSize=10, textColor=BLEU_FONCE)),
            Paragraph(
                f'{salle_nom} — {niveau_nom}',
                s('fv', fontSize=10)
            ),
        ],
    ], colWidths=[5.5*cm, 12.5*cm])

    fiche.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), GRIS_CLAIR),
        ('BACKGROUND', (0,0), (0,-1), BLEU_CLAIR),
        ('GRID', (0,0), (-1,-1), 0.3, BORDURE),
        ('TOPPADDING', (0,0), (-1,-1), 7),
        ('BOTTOMPADDING', (0,0), (-1,-1), 7),
        ('LEFTPADDING', (0,0), (-1,-1), 10),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BACKGROUND', (0,2), (-1,2), colors.HexColor('#EFF6FF')),
    ]))
    elements.append(fiche)
    elements.append(Spacer(1, 0.5*cm))

    elements.append(Paragraph(
        f"{article} est bien {inscrit}(e) dans notre établissement "
        f"et fréquente régulièrement les cours durant l'année scolaire "
        f"<b>{annee.nom}</b> en classe de <b>{salle_nom}</b>.<br/><br/>"
        f"Le présent certificat est délivré à l'intéressé(e) pour "
        f"servir et valoir ce que de droit "
        f"(<i>{certificat.motif}</i>).",
        s('fin', fontSize=11, leading=18, alignment=TA_JUSTIFY)
    ))
    elements.append(Spacer(1, 0.8*cm))

    # ── DATE ──────────────────────────────────────────────────────────────────
    from django.utils import timezone
    date_str = certificat.date_delivrance.strftime('%d/%m/%Y')
    elements.append(Paragraph(
        f"Fait à Lomé, le {date_str}",
        s('date', fontSize=10, alignment=TA_CENTER)
    ))
    elements.append(Spacer(1, 0.4*cm))

    # ── SIGNATURES + QR ───────────────────────────────────────────────────────
    sig = Table([[
        Paragraph(
            "Le Directeur<br/><br/><br/>"
            "(Signature et cachet)",
            s('sig', fontSize=9, alignment=TA_CENTER,
              textColor=colors.HexColor('#64748B'))
        ),
        Table([[
            Paragraph(
                f"<b>N° {certificat.numero}</b>",
                s('num', fontSize=8, alignment=TA_CENTER,
                  textColor=BLEU_FONCE)
            )], [qr_image]],
            colWidths=[3*cm]
        ),
    ]], colWidths=[16*cm, 3*cm])
    sig.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'BOTTOM'),
        ('ALIGN', (1,0), (1,-1), 'CENTER'),
        ('BOX', (0,0), (0,-1), 0.5, BORDURE),
        ('TOPPADDING', (0,0), (-1,-1), 8),
    ]))
    elements.append(sig)

    # ── PIED DE PAGE ──────────────────────────────────────────────────────────
    elements.append(Spacer(1, 0.4*cm))
    elements.append(HRFlowable(
        width='100%', thickness=1, color=BLEU_FONCE
    ))
    elements.append(Paragraph(
        f"{config.nom} | {config.adresse or ''} | "
        f"Tél : {config.telephone or ''} | {config.email or ''}",
        s('pied', fontSize=7, alignment=TA_CENTER,
          textColor=colors.HexColor('#94A3B8'), spaceBefore=4)
    ))

    doc.build(
        elements,
        onFirstPage=filigrane,
        onLaterPages=filigrane
    )
    buffer.seek(0)

    response = HttpResponse(content_type='application/pdf')
    nom_fichier = (
        f"certificat_{eleve.matricule}_{annee.nom}.pdf"
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{nom_fichier}"'
    )
    response.write(buffer.getvalue())
    return response


# ── ATTESTATION DE FREQUENTATION ─────────────────────────────────────────────

@login_required
@role_requis('DIRECTEUR', 'CENSEUR', 'SECRETAIRE')
def attestation_pdf(request, eleve_pk):
    eleve = get_object_or_404(Eleve, pk=eleve_pk)
    annee = AnneeScolaire.active()
    config = ConfigurationDocument.get()

    if not annee:
        messages.error(request, "Aucune annee active.")
        return redirect('liste_documents')

    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY

    buffer = io.BytesIO()
    BLEU = colors.HexColor('#1E40AF')
    styles = getSampleStyleSheet()

    titre_style = ParagraphStyle(
        't', parent=styles['Normal'],
        fontSize=15, fontName='Helvetica-Bold',
        textColor=BLEU, alignment=TA_CENTER,
    )
    normal = ParagraphStyle(
        'n', parent=styles['Normal'], fontSize=10, leading=16,
    )
    justifie = ParagraphStyle(
        'j', parent=styles['Normal'],
        fontSize=11, leading=18, alignment=TA_JUSTIFY,
    )
    centre = ParagraphStyle(
        'c', parent=styles['Normal'],
        fontSize=10, alignment=TA_CENTER,
    )

    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )

    inscription = eleve.inscriptions.filter(
        annee=annee, statut='ACTIVE'
    ).select_related('salle__niveau').first()

    elements = []

    elements.append(Paragraph(
        f"<b>{config.nom_ecole}</b><br/>"
        f"{config.adresse or ''} — Tel: {config.telephone or ''}",
        ParagraphStyle('h', parent=styles['Normal'],
                      fontSize=10, alignment=TA_CENTER)
    ))
    elements.append(Spacer(1, 0.4*cm))
    elements.append(Paragraph("ATTESTATION DE FREQUENTATION", titre_style))
    elements.append(Spacer(1, 0.6*cm))

    salle_nom = inscription.salle.nom if inscription else 'Non inscrit'
    sexe = 'la nommee' if eleve.sexe == 'F' else 'le nomme'
    article = 'Elle' if eleve.sexe == 'F' else 'Il'
    inscrit = 'inscrite' if eleve.sexe == 'F' else 'inscrit'

    corps = (
        f"Je soussigne, Directeur de <b>{config.nom_ecole}</b>, "
        f"atteste que {sexe} <b>{eleve.nom_complet}</b>, "
        f"matricule <b>{eleve.matricule}</b>, "
        f"est bien {inscrit}(e) et frequente notre etablissement "
        f"en classe de <b>{salle_nom}</b> "
        f"pour l'annee scolaire <b>{annee.nom}</b>.<br/><br/>"
        f"La presente attestation est delivree a l'interesse(e) "
        f"pour servir et valoir ce que de droit."
    )
    elements.append(Paragraph(corps, justifie))
    elements.append(Spacer(1, 1*cm))

    from django.utils import timezone
    date_str = timezone.now().date().strftime('%d/%m/%Y')
    elements.append(Paragraph(f"Fait a Lome, le {date_str}", centre))
    elements.append(Spacer(1, 0.4*cm))

    sig_data = [[
        Paragraph(
            "Le Directeur<br/><br/><br/>(Signature et cachet)",
            centre
        ),
    ]]
    sig_table = Table(sig_data, colWidths=[18*cm])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 8),
    ]))
    elements.append(sig_table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(content_type='application/pdf')
    nom = f"attestation_{eleve.matricule}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{nom}"'
    response.write(buffer.getvalue())
    return response


# ── EXPORT LISTE ELEVES PDF ───────────────────────────────────────────────────

@login_required
@role_requis('DIRECTEUR', 'CENSEUR', 'SECRETAIRE')
def export_liste_eleves_pdf(request):
    annee = AnneeScolaire.active()
    salle_pk = request.GET.get('salle', '')
    config = ConfigurationDocument.get()

    inscriptions = Inscription.objects.filter(
        annee=annee, statut='ACTIVE'
    ).select_related(
        'eleve', 'salle', 'salle__niveau'
    ).order_by('salle__niveau__ordre', 'salle__nom', 'eleve__nom')

    if salle_pk:
        inscriptions = inscriptions.filter(salle__pk=salle_pk)

    import io
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_CENTER

    buffer = io.BytesIO()
    BLEU = colors.HexColor('#1E40AF')
    styles = getSampleStyleSheet()

    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )

    elements = []

    titre = ParagraphStyle(
        't', parent=styles['Normal'],
        fontSize=13, fontName='Helvetica-Bold',
        textColor=BLEU, alignment=TA_CENTER,
    )
    petit = ParagraphStyle(
        'p', parent=styles['Normal'], fontSize=8,
    )

    elements.append(Paragraph(
        f"{config.nom_ecole} — Liste des eleves inscrits — "
        f"{annee.nom if annee else ''}",
        titre
    ))
    elements.append(Spacer(1, 0.3*cm))

    header = [
        Paragraph('<b>N°</b>', ParagraphStyle('h', parent=styles['Normal'],
                  fontSize=8, fontName='Helvetica-Bold',
                  alignment=TA_CENTER)),
        Paragraph('<b>Matricule</b>', ParagraphStyle('h', parent=styles['Normal'],
                  fontSize=8, fontName='Helvetica-Bold')),
        Paragraph('<b>Nom complet</b>', ParagraphStyle('h', parent=styles['Normal'],
                  fontSize=8, fontName='Helvetica-Bold')),
        Paragraph('<b>Sexe</b>', ParagraphStyle('h', parent=styles['Normal'],
                  fontSize=8, fontName='Helvetica-Bold', alignment=TA_CENTER)),
        Paragraph('<b>Date naiss.</b>', ParagraphStyle('h', parent=styles['Normal'],
                  fontSize=8, fontName='Helvetica-Bold')),
        Paragraph('<b>Classe</b>', ParagraphStyle('h', parent=styles['Normal'],
                  fontSize=8, fontName='Helvetica-Bold')),
        Paragraph('<b>Niveau</b>', ParagraphStyle('h', parent=styles['Normal'],
                  fontSize=8, fontName='Helvetica-Bold')),
        Paragraph('<b>Obs.</b>', ParagraphStyle('h', parent=styles['Normal'],
                  fontSize=8, fontName='Helvetica-Bold')),
    ]

    data = [header]
    for i, insc in enumerate(inscriptions, 1):
        e = insc.eleve
        data.append([
            Paragraph(str(i), ParagraphStyle('c', parent=styles['Normal'],
                      fontSize=7, alignment=TA_CENTER)),
            Paragraph(e.matricule, petit),
            Paragraph(e.nom_complet, petit),
            Paragraph(e.get_sexe_display(), ParagraphStyle('c',
                      parent=styles['Normal'], fontSize=7,
                      alignment=TA_CENTER)),
            Paragraph(
                e.date_naissance.strftime('%d/%m/%Y')
                if e.date_naissance else '—',
                petit
            ),
            Paragraph(insc.salle.nom, petit),
            Paragraph(insc.salle.niveau.nom, petit),
            Paragraph(
                'Redoublant' if e.redoublant else '',
                ParagraphStyle('c', parent=styles['Normal'], fontSize=7)
            ),
        ])

    col_widths = [
        1*cm, 2.5*cm, 5*cm, 1.2*cm, 2.5*cm, 2*cm, 2*cm, 2.5*cm
    ]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), BLEU),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [
            colors.white, colors.HexColor('#F8FAFC')
        ]),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.3*cm))
    elements.append(Paragraph(
        f"Total : {len(data)-1} eleve(s)",
        ParagraphStyle('f', parent=styles['Normal'],
                      fontSize=9, textColor=BLEU)
    ))

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="liste_eleves_{annee.nom if annee else ""}.pdf"'
    )
    response.write(buffer.getvalue())
    return response


# ── EXPORT LISTE ELEVES EXCEL ─────────────────────────────────────────────────

@login_required
@role_requis('DIRECTEUR', 'CENSEUR', 'SECRETAIRE')
def export_liste_eleves_excel(request):
    annee = AnneeScolaire.active()
    salle_pk = request.GET.get('salle', '')

    inscriptions = Inscription.objects.filter(
        annee=annee, statut='ACTIVE'
    ).select_related(
        'eleve', 'salle', 'salle__niveau'
    ).order_by('salle__niveau__ordre', 'salle__nom', 'eleve__nom')

    if salle_pk:
        inscriptions = inscriptions.filter(salle__pk=salle_pk)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Eleves"

    BLEU = "1E40AF"
    bleu_fill = PatternFill(fill_type="solid", fgColor=BLEU)
    blanc_font = Font(bold=True, color="FFFFFF", size=10)
    bold_font = Font(bold=True, size=10)
    thin = Side(border_style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        'N°', 'Matricule', 'Nom', 'Prenom', 'Sexe',
        'Date naissance', 'Lieu naissance', 'Classe',
        'Niveau', 'Redoublant', 'Contact urgence', 'Tel urgence',
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = blanc_font
        cell.fill = bleu_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    ws.row_dimensions[1].height = 20

    for row, insc in enumerate(inscriptions, 2):
        e = insc.eleve
        valeurs = [
            row - 1,
            e.matricule,
            e.nom,
            e.prenom,
            e.get_sexe_display(),
            e.date_naissance.strftime('%d/%m/%Y') if e.date_naissance else '',
            e.lieu_naissance,
            insc.salle.nom,
            insc.salle.niveau.nom,
            'Oui' if e.redoublant else 'Non',
            e.contact_urgence,
            e.telephone_urgence,
        ]
        for col, val in enumerate(valeurs, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if row % 2 == 0:
                cell.fill = PatternFill(
                    fill_type="solid", fgColor="F8FAFC"
                )

    widths = [5, 15, 15, 15, 8, 14, 16, 10, 12, 12, 20, 16]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col)
        ].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument'
                     '.spreadsheetml.sheet'
    )
    nom = f"eleves_{annee.nom if annee else 'export'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nom}"'
    wb.save(response)
    return response


# ── BILAN ANNUEL PDF ──────────────────────────────────────────────────────────

@login_required
@role_requis('DIRECTEUR', 'CENSEUR')
def bilan_annuel_pdf(request):
    annee = AnneeScolaire.active()
    config = ConfigurationDocument.get()

    if not annee:
        messages.error(request, "Aucune annee active.")
        return redirect('liste_documents')

    from apps.grades.models import MoyenneGenerale
    from apps.finance.models import Paiement, FraisEleve
    from apps.students.models import Inscription
    from django.db.models import Sum, Avg, Count

    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_CENTER

    buffer = io.BytesIO()
    BLEU = colors.HexColor('#1E40AF')
    styles = getSampleStyleSheet()
    titre_s = ParagraphStyle(
        't', parent=styles['Normal'],
        fontSize=14, fontName='Helvetica-Bold',
        textColor=BLEU, alignment=TA_CENTER,
    )
    section_s = ParagraphStyle(
        's', parent=styles['Normal'],
        fontSize=11, fontName='Helvetica-Bold',
        textColor=BLEU, spaceBefore=10,
    )
    normal_s = ParagraphStyle(
        'n', parent=styles['Normal'], fontSize=9,
    )

    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )

    elements = []
    elements.append(Paragraph(
        f"{config.nom_ecole}", titre_s
    ))
    elements.append(Paragraph(
        f"BILAN ANNUEL — {annee.nom}",
        ParagraphStyle('b', parent=styles['Normal'],
                      fontSize=12, fontName='Helvetica-Bold',
                      alignment=TA_CENTER, spaceAfter=10)
    ))
    elements.append(Spacer(1, 0.3*cm))

    # Section 1 — Effectifs
    elements.append(Paragraph("1. EFFECTIFS", section_s))
    nb_eleves = Inscription.objects.filter(
        annee=annee, statut='ACTIVE'
    ).count()
    nb_garcons = Inscription.objects.filter(
        annee=annee, statut='ACTIVE', eleve__sexe='M'
    ).count()
    nb_filles = Inscription.objects.filter(
        annee=annee, statut='ACTIVE', eleve__sexe='F'
    ).count()
    nb_redoublants = Inscription.objects.filter(
        annee=annee, statut='ACTIVE', eleve__redoublant=True
    ).count()

    eff_data = [
        ['Total eleves inscrits', str(nb_eleves)],
        ['Garcons', str(nb_garcons)],
        ['Filles', str(nb_filles)],
        ['Redoublants', str(nb_redoublants)],
    ]
    eff_table = Table(eff_data, colWidths=[12*cm, 6*cm])
    eff_table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [
            colors.white, colors.HexColor('#F8FAFC')
        ]),
        ('FONTNAME', (1,0), (1,-1), 'Helvetica-Bold'),
        ('ALIGN', (1,0), (1,-1), 'CENTER'),
    ]))
    elements.append(eff_table)

    # Section 2 — Resultats par niveau
    elements.append(Paragraph("2. RESULTATS PAR NIVEAU", section_s))
    niveaux = Niveau.objects.all().order_by('ordre')
    res_header = [
        'Niveau', 'Effectif', 'Admis', 'Echec', 'Taux'
    ]
    res_data = [res_header]
    for niveau in niveaux:
        inscrits = Inscription.objects.filter(
            salle__niveau=niveau, annee=annee, statut='ACTIVE'
        ).count()
        if inscrits == 0:
            continue
        admis = MoyenneGenerale.objects.filter(
            eleve__inscriptions__salle__niveau=niveau,
            eleve__inscriptions__annee=annee,
            decision='ADMIS',
        ).count()
        taux = round(admis / inscrits * 100, 1) if inscrits > 0 else 0
        res_data.append([
            niveau.nom,
            str(inscrits),
            str(admis),
            str(inscrits - admis),
            f"{taux}%",
        ])

    if len(res_data) > 1:
        res_table = Table(
            res_data,
            colWidths=[5*cm, 3*cm, 3*cm, 3*cm, 4*cm]
        )
        res_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), BLEU),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#E2E8F0')),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('ALIGN', (1,0), (-1,-1), 'CENTER'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [
                colors.white, colors.HexColor('#F8FAFC')
            ]),
        ]))
        elements.append(res_table)

    # Section 3 — Finance
    elements.append(Paragraph("3. SITUATION FINANCIERE", section_s))
    total_recettes = float(
        Paiement.objects.aggregate(t=Sum('montant'))['t'] or 0
    )
    frais = FraisEleve.objects.filter(annee=annee)
    total_du = float(frais.aggregate(t=Sum('montant'))['t'] or 0)
    total_paye = float(frais.aggregate(t=Sum('montant_paye'))['t'] or 0)
    taux_rec = round(total_paye / total_du * 100, 1) if total_du > 0 else 0

    fin_data = [
        ['Total recettes encaissees', f"{total_recettes:,.0f} FCFA"],
        ['Total frais attendus', f"{total_du:,.0f} FCFA"],
        ['Total frais payes', f"{total_paye:,.0f} FCFA"],
        ['Total impayes', f"{total_du - total_paye:,.0f} FCFA"],
        ['Taux de recouvrement', f"{taux_rec}%"],
    ]
    fin_table = Table(fin_data, colWidths=[12*cm, 6*cm])
    fin_table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [
            colors.white, colors.HexColor('#F8FAFC')
        ]),
        ('FONTNAME', (1,0), (1,-1), 'Helvetica-Bold'),
        ('ALIGN', (1,0), (1,-1), 'CENTER'),
    ]))
    elements.append(fin_table)
    elements.append(Spacer(1, 0.5*cm))

    from django.utils import timezone
    elements.append(Paragraph(
        f"Rapport genere le {timezone.now().strftime('%d/%m/%Y a %H:%M')}",
        ParagraphStyle('f', parent=styles['Normal'],
                      fontSize=8, textColor=colors.HexColor('#64748B'),
                      alignment=TA_CENTER)
    ))

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="bilan_{annee.nom}.pdf"'
    )
    response.write(buffer.getvalue())
    return response
