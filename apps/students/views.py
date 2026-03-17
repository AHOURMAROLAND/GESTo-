from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import models
from .models import Eleve, Parent, EleveParent, Inscription
from apps.academic.models import SalleClasse, AnneeScolaire, MatiereSalle
from apps.authentication.models import CustomUser


def role_requis(*roles):
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('login')
            if not (request.user.is_superuser or request.user.role in roles):
                messages.error(request, "Acces refuse.")
                return redirect('dashboard')
            return view_func(request, *args, **kwargs)
        wrapper.__name__ = view_func.__name__
        return wrapper
    return decorator


def generer_matricule(annee):
    """Genere un matricule unique : BKT-2025-0001"""
    from django.utils import timezone
    annee_str = str(timezone.now().year)
    prefix = f"BKT-{annee_str}-"
    derniers = Eleve.objects.filter(
        matricule__startswith=prefix
    ).order_by('-matricule')
    if derniers.exists():
        dernier = derniers.first().matricule
        num = int(dernier.split('-')[-1]) + 1
    else:
        num = 1
    return f"{prefix}{num:04d}"


# ── ELEVES ────────────────────────────────────────────────────────────────────

@login_required
def liste_eleves(request):
    annee = AnneeScolaire.active()
    q = request.GET.get('q', '')
    salle_pk = request.GET.get('salle', '')
    statut_f = request.GET.get('statut', 'ACTIVE')

    # Base queryset
    inscriptions = Inscription.objects.filter(
        annee=annee
    ).select_related(
        'eleve', 'salle', 'salle__niveau'
    ).order_by(
        'salle__niveau__ordre',
        'salle__nom',
        'eleve__nom',
        'eleve__prenom',
    ) if annee else Inscription.objects.none()

    if q:
        inscriptions = inscriptions.filter(
            models.Q(eleve__nom__icontains=q) |
            models.Q(eleve__prenom__icontains=q) |
            models.Q(eleve__matricule__icontains=q)
        )

    if salle_pk:
        inscriptions = inscriptions.filter(salle__pk=salle_pk)

    if statut_f:
        inscriptions = inscriptions.filter(statut=statut_f)

    # Restriction PROFESSEUR
    if request.user.role == 'PROFESSEUR':
        salles_prof = MatiereSalle.objects.filter(
            professeur=request.user, salle__annee=annee
        ).values_list('salle_id', flat=True)
        inscriptions = inscriptions.filter(salle__pk__in=salles_prof)

    # Grouper par salle
    # Recompter correctement
    groupes_salles = []
    salle_courante = None
    inscrits_courants = []

    for insc in inscriptions:
        if salle_courante is None:
            salle_courante = insc.salle
        if insc.salle != salle_courante:
            groupes_salles.append({
                'salle': salle_courante,
                'inscriptions': inscrits_courants,
                'nb': len(inscrits_courants),
            })
            salle_courante = insc.salle
            inscrits_courants = []
        inscrits_courants.append(insc)

    if salle_courante and inscrits_courants:
        groupes_salles.append({
            'salle': salle_courante,
            'inscriptions': inscrits_courants,
            'nb': len(inscrits_courants),
        })

    salles = SalleClasse.objects.filter(
        annee=annee, est_active=True
    ).order_by('niveau__ordre', 'nom') if annee else []

    total = inscriptions.count()

    return render(request, 'students/liste_eleves.html', {
        'groupes_salles': groupes_salles,
        'inscriptions': inscriptions,
        'salles': salles,
        'q': q,
        'salle_pk': salle_pk,
        'statut_filtre': statut_f,
        'total': total,
        'annee': annee,
    })


@login_required
@role_requis('DIRECTEUR', 'CENSEUR', 'SECRETAIRE')
def nouvel_eleve(request):
    annee = AnneeScolaire.active()
    if not annee:
        messages.error(request, "Activez d'abord une annee scolaire.")
        return redirect('liste_eleves')

    salles = SalleClasse.objects.filter(
        annee=annee, est_active=True
    ).order_by('niveau__ordre', 'nom')

    if request.method == 'POST':
        nom = request.POST.get('nom', '').strip().upper()
        prenom = request.POST.get('prenom', '').strip()
        sexe = request.POST.get('sexe', 'M')
        date_naissance = request.POST.get('date_naissance') or None
        lieu_naissance = request.POST.get('lieu_naissance', '').strip()
        salle_pk = request.POST.get('salle_id')
        redoublant = request.POST.get('redoublant') == '1'
        contact_urgence = request.POST.get('contact_urgence', '').strip()
        telephone_urgence = request.POST.get('telephone_urgence', '').strip()

        if not all([nom, prenom, salle_pk]):
            messages.error(request, "Nom, prenom et salle sont obligatoires.")
            return render(request, 'students/nouvel_eleve.html', {
                'salles': salles, 'annee': annee
            })

        matricule = generer_matricule(annee)

        # Creer compte utilisateur pour l'eleve
        username = f"{prenom[:6].lower()}.{nom[:4].lower()}".replace(' ', '')
        base = username
        cpt = 1
        while CustomUser.objects.filter(username=username).exists():
            username = f"{base}{cpt}"
            cpt += 1

        import random
        pwd = f"bkt{random.randint(1000, 9999)}"

        from django.contrib.auth.hashers import make_password
        user = CustomUser.objects.create(
            username=username,
            first_name=prenom,
            last_name=nom,
            role='ELEVE',
            password=make_password(pwd),
        )

        eleve = Eleve.objects.create(
            user=user,
            matricule=matricule,
            nom=nom,
            prenom=prenom,
            sexe=sexe,
            date_naissance=date_naissance,
            lieu_naissance=lieu_naissance,
            redoublant=redoublant,
            contact_urgence=contact_urgence,
            telephone_urgence=telephone_urgence,
            statut='ACTIF',
        )

        if request.FILES.get('photo'):
            eleve.photo = request.FILES['photo']
            eleve.save()

        Inscription.objects.create(
            eleve=eleve,
            salle_id=salle_pk,
            annee=annee,
            statut='ACTIVE',
        )

        messages.success(
            request,
            f"Eleve {prenom} {nom} inscrit — Matricule: {matricule} "
            f"| Login: {username} | MDP: {pwd}"
        )
        return redirect('detail_eleve', pk=eleve.pk)

    return render(request, 'students/nouvel_eleve.html', {
        'salles': salles,
        'annee': annee,
    })


@login_required
def detail_eleve(request, pk):
    eleve = get_object_or_404(Eleve, pk=pk)
    inscriptions = eleve.inscriptions.select_related(
        'salle', 'salle__niveau', 'annee'
    ).order_by('-annee__nom')
    parents = EleveParent.objects.filter(
        eleve=eleve
    ).select_related('parent')
    inscription_active = eleve.inscription_active

    from apps.devoirs.models import SoumissionDevoir
    soumissions_devoirs = SoumissionDevoir.objects.filter(
        eleve=eleve
    ).select_related('devoir', 'devoir__matiere_salle__matiere').order_by('-date_soumission')[:5]

    return render(request, 'students/detail_eleve.html', {
        'eleve': eleve,
        'inscriptions': inscriptions,
        'parents': parents,
        'inscription_active': inscription_active,
        'soumissions_devoirs': soumissions_devoirs,
    })


@login_required
@role_requis('DIRECTEUR', 'CENSEUR', 'SECRETAIRE')
def modifier_eleve(request, pk):
    eleve = get_object_or_404(Eleve, pk=pk)
    if request.method == 'POST':
        eleve.nom = request.POST.get('nom', '').strip().upper()
        eleve.prenom = request.POST.get('prenom', '').strip()
        eleve.sexe = request.POST.get('sexe', 'M')
        eleve.date_naissance = request.POST.get('date_naissance') or None
        eleve.lieu_naissance = request.POST.get('lieu_naissance', '').strip()
        eleve.contact_urgence = request.POST.get('contact_urgence', '').strip()
        eleve.telephone_urgence = request.POST.get('telephone_urgence', '').strip()
        eleve.groupe_sanguin = request.POST.get('groupe_sanguin', '').strip()
        eleve.allergies = request.POST.get('allergies', '').strip()
        eleve.statut = request.POST.get('statut', 'ACTIF')
        eleve.redoublant = request.POST.get('redoublant') == '1'
        if request.FILES.get('photo'):
            eleve.photo = request.FILES['photo']
        eleve.save()
        messages.success(request, "Fiche eleve mise a jour.")
        return redirect('detail_eleve', pk=pk)

    return render(request, 'students/modifier_eleve.html', {'eleve': eleve})


@login_required
@role_requis('DIRECTEUR', 'CENSEUR', 'SECRETAIRE')
def transferer_eleve(request, pk):
    eleve = get_object_or_404(Eleve, pk=pk)
    annee = AnneeScolaire.active()
    salles = SalleClasse.objects.filter(
        annee=annee, est_active=True
    ).order_by('niveau__ordre', 'nom')
    inscription = eleve.inscription_active

    if request.method == 'POST':
        nouvelle_salle_pk = request.POST.get('salle_id')
        if not nouvelle_salle_pk:
            messages.error(request, "Choisissez une salle.")
            return redirect('transferer_eleve', pk=pk)

        if inscription:
            ancienne = inscription.salle.nom
            inscription.salle_id = nouvelle_salle_pk
            inscription.save()
            messages.success(
                request,
                f"{eleve.nom_complet} transfere de {ancienne} "
                f"vers {inscription.salle.nom}."
            )
        return redirect('detail_eleve', pk=pk)

    return render(request, 'students/transferer_eleve.html', {
        'eleve': eleve,
        'salles': salles,
        'inscription': inscription,
    })


# ── PARENTS ───────────────────────────────────────────────────────────────────

@login_required
@role_requis('DIRECTEUR', 'CENSEUR', 'SECRETAIRE')
def ajouter_parent(request, eleve_pk):
    eleve = get_object_or_404(Eleve, pk=eleve_pk)

    if request.method == 'POST':
        action = request.POST.get('action', 'nouveau')

        if action == 'lier':
            # Lier un parent existant
            parent_pk = request.POST.get('parent_id')
            lien = request.POST.get('lien', 'TUTEUR')
            est_principal = request.POST.get('est_principal') == '1'
            if parent_pk:
                if EleveParent.objects.filter(
                    eleve=eleve, parent_id=parent_pk
                ).exists():
                    messages.error(request, "Ce parent est deja lie a cet eleve.")
                else:
                    EleveParent.objects.create(
                        eleve=eleve,
                        parent_id=parent_pk,
                        lien=lien,
                        est_contact_principal=est_principal,
                    )
                    messages.success(request, "Parent lie.")
        else:
            # Creer nouveau parent
            nom = request.POST.get('nom', '').strip().upper()
            prenom = request.POST.get('prenom', '').strip()
            telephone = request.POST.get('telephone', '').strip()
            telephone_wa = request.POST.get('telephone_wa', '').strip()
            email = request.POST.get('email', '').strip()
            profession = request.POST.get('profession', '').strip()
            adresse = request.POST.get('adresse', '').strip()
            lien = request.POST.get('lien', 'TUTEUR')
            est_principal = request.POST.get('est_principal') == '1'
            langue = request.POST.get('langue', 'FR')

            if not nom:
                messages.error(request, "Le nom est obligatoire.")
                return redirect('detail_eleve', pk=eleve_pk)

            # Creer compte utilisateur parent
            username_base = f"parent.{nom[:6].lower()}"
            username = username_base.replace(' ', '')
            cpt = 1
            while CustomUser.objects.filter(username=username).exists():
                username = f"{username_base}{cpt}"
                cpt += 1

            import random
            from django.contrib.auth.hashers import make_password
            pwd = f"bkt{random.randint(1000, 9999)}"

            user = CustomUser.objects.create(
                username=username,
                first_name=prenom,
                last_name=nom,
                role='PARENT',
                telephone=telephone,
                telephone_wa=telephone_wa,
                password=make_password(pwd),
            )

            parent = Parent.objects.create(
                user=user,
                nom=nom,
                prenom=prenom,
                telephone=telephone,
                telephone_wa=telephone_wa,
                email=email,
                profession=profession,
                adresse=adresse,
                langue=langue,
            )

            EleveParent.objects.create(
                eleve=eleve,
                parent=parent,
                lien=lien,
                est_contact_principal=est_principal,
            )

            messages.success(
                request,
                f"Parent {prenom} {nom} cree — "
                f"Login: {username} | MDP: {pwd}"
            )

        return redirect('detail_eleve', pk=eleve_pk)

    parents_existants = Parent.objects.all().order_by('nom')
    return render(request, 'students/ajouter_parent.html', {
        'eleve': eleve,
        'parents_existants': parents_existants,
        'liens': EleveParent._meta.get_field('lien').choices,
        'langues': Parent._meta.get_field('langue').choices,
    })


@login_required
@role_requis('DIRECTEUR', 'CENSEUR', 'SECRETAIRE')
def retirer_parent(request, eleve_pk, parent_pk):
    if request.method == 'POST':
        ep = get_object_or_404(EleveParent, eleve_id=eleve_pk, parent_id=parent_pk)
        ep.delete()
        messages.success(request, "Lien parent retire.")
    return redirect('detail_eleve', pk=eleve_pk)


# ── EXPORT EXCEL ──────────────────────────────────────────────────────────────

@login_required
@role_requis('DIRECTEUR', 'CENSEUR', 'SECRETAIRE')
def export_eleves_excel(request):
    import openpyxl
    from django.http import HttpResponse

    annee = AnneeScolaire.active()
    salle_pk = request.GET.get('salle', '')

    inscriptions = Inscription.objects.select_related(
        'eleve', 'salle', 'salle__niveau', 'annee'
    ).filter(annee=annee, statut='ACTIVE')

    if salle_pk:
        inscriptions = inscriptions.filter(salle__pk=salle_pk)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Eleves"

    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1E40AF")

    headers = [
        'Matricule', 'Nom', 'Prenom', 'Sexe',
        'Date naissance', 'Lieu naissance',
        'Salle', 'Niveau', 'Statut', 'Redoublant'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, insc in enumerate(inscriptions, 2):
        e = insc.eleve
        ws.cell(row=row, column=1, value=e.matricule)
        ws.cell(row=row, column=2, value=e.nom)
        ws.cell(row=row, column=3, value=e.prenom)
        ws.cell(row=row, column=4, value=e.get_sexe_display())
        ws.cell(row=row, column=5,
                value=str(e.date_naissance) if e.date_naissance else '')
        ws.cell(row=row, column=6, value=e.lieu_naissance)
        ws.cell(row=row, column=7, value=insc.salle.nom)
        ws.cell(row=row, column=8, value=insc.salle.niveau.nom)
        ws.cell(row=row, column=9, value=insc.get_statut_display())
        ws.cell(row=row, column=10, value='Oui' if e.redoublant else 'Non')

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument'
                      '.spreadsheetml.sheet'
    )
    nom_fichier = f"eleves_{annee.nom if annee else 'export'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nom_fichier}"'
    wb.save(response)
    return response


# ── VALIDATION NUMERO WA ──────────────────────────────────────────────────────

@login_required
def verifier_numero_wa(request):
    import re
    import json
    from django.http import JsonResponse

    numero = request.GET.get('numero', '').strip().replace(' ', '').replace('-', '')
    if not numero.startswith('+'):
        numero = '+' + numero

    if not re.match(r'^\+\d{10,15}$', numero):
        return JsonResponse({
            'valide': False,
            'message': 'Format invalide. Utilisez +228XXXXXXXX'
        })

    from django.conf import settings
    wa_key = getattr(settings, 'WA_API_KEY', '')

    if not wa_key:
        return JsonResponse({
            'valide': True,
            'message': 'Non verifie (API non configuree)',
            'non_verifie': True,
        })

    try:
        import requests as req
        r = req.post(
            f"{settings.WA_BASE_URL}/api/checkWhatsapp",
            json={'phoneNumber': numero},
            headers={'Authorization': f'Bearer {wa_key}'},
            timeout=5
        )
        data = r.json()
        existe = data.get('existsWhatsapp', data.get('exists', False))
        return JsonResponse({
            'valide': existe,
            'message': (
                'Numero WhatsApp verifie' if existe
                else 'Ce numero n\'a pas WhatsApp'
            )
        })
    except Exception:
        return JsonResponse({
            'valide': True,
            'message': 'Non verifie (hors ligne)',
            'non_verifie': True,
        })


# ── EXPORT PDF ────────────────────────────────────────────────────────────────

@login_required
@role_requis('DIRECTEUR', 'CENSEUR', 'SECRETAIRE')
def export_eleves_pdf(request):
    """Export PDF de la liste des eleves."""
    from django.http import HttpResponse
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    import io

    annee = AnneeScolaire.active()
    salle_pk = request.GET.get('salle', '')

    inscriptions = Inscription.objects.select_related(
        'eleve', 'salle', 'salle__niveau'
    ).filter(annee=annee, statut='ACTIVE').order_by(
        'salle__niveau__ordre', 'salle__nom', 'eleve__nom'
    )

    if salle_pk:
        inscriptions = inscriptions.filter(salle__pk=salle_pk)

    buffer = io.BytesIO()
    BLEU = colors.HexColor('#1E3A8A')
    GRIS = colors.HexColor('#F8FAFC')
    styles = getSampleStyleSheet()

    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                             leftMargin=1.5*cm, rightMargin=1.5*cm,
                             topMargin=2*cm, bottomMargin=1.5*cm)

    elements = []
    elements.append(Paragraph(
        f"Liste des Eleves — {annee.nom if annee else ''}",
        ParagraphStyle('t', parent=styles['Normal'], fontSize=14,
                       fontName='Helvetica-Bold', textColor=BLEU,
                       alignment=1, spaceAfter=12)
    ))

    header = ['#', 'Matricule', 'Nom et Prenoms', 'Sexe', 'Date Naiss.', 'Classe', 'Niveau', 'Statut']
    data = [header]
    for i, insc in enumerate(inscriptions, 1):
        e = insc.eleve
        data.append([
            str(i), e.matricule, e.nom_complet,
            e.get_sexe_display(),
            e.date_naissance.strftime('%d/%m/%Y') if e.date_naissance else '-',
            insc.salle.nom, insc.salle.niveau.nom, insc.get_statut_display()
        ])

    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), BLEU),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, GRIS]),
    ]))
    elements.append(t)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="eleves_{annee.nom if annee else "export"}.pdf"'
    response.write(buffer.getvalue())
    return response


# ── MES ENFANTS (PARENT) ─────────────────────────────────────────────────────

@login_required
def mes_enfants(request):
    """Vue parent: affiche ses enfants."""
    from .models import EleveParent
    try:
        parent = request.user.profil_parent
        eleves_parents = EleveParent.objects.filter(
            parent=parent
        ).select_related('eleve')
    except Exception:
        eleves_parents = []

    return render(request, 'students/mes_enfants.html', {
        'eleves_parents': eleves_parents,
    })
